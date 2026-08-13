#!/usr/bin/env python
from __future__ import annotations

import argparse
import json
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import compare_runs as cr
from computational_complexity import collect_solver_log_detail, read_scalar_file, summarize_solver_logs

PROJECT_ROOT = Path(__file__).resolve().parents[1]
METHOD_CHOICES = ("monolithic", "benders", "lbbd")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Compare two or more completed runs produced by the same optimization method "
            "across technology/redirection scenarios."
        )
    )
    parser.add_argument("--method", choices=METHOD_CHOICES, required=True)
    parser.add_argument("--run", action="append", required=True, help="Run folder; repeat for every scenario.")
    parser.add_argument("--label", action="append", default=[], help="Optional label; repeat in run order.")
    parser.add_argument(
        "--baseline-index", type=int, default=1,
        help="1-based index of the baseline run (default: first run).",
    )
    parser.add_argument("--out", default=None, help="Optional output XLSX path.")
    parser.add_argument("--project-root", default=str(PROJECT_ROOT))
    return parser.parse_args()


def _resolve(root: Path, value: str) -> Path:
    path = Path(value)
    return path.resolve() if path.is_absolute() else (root / path).resolve()


def _read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _read_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def _detect_method(run_dir: Path) -> str:
    metadata = _read_json(run_dir / "run_metadata.json")
    benders = _read_json(run_dir / "logs" / "benders_manifest.json")
    for value in (metadata.get("method"), benders.get("method")):
        if value:
            text = str(value).strip().lower()
            if "benders" in text and "logic" not in text and text != "lbbd":
                return "benders"
            if text == "lbbd" or "logic" in text:
                return "lbbd"
            if "monolithic" in text:
                return "monolithic"
    probe = f"{run_dir.name}\n{_read_text(run_dir / 'README_RUN.txt')[:2500]}".lower()
    if "lbbd" in probe:
        return "lbbd"
    if "benders" in probe:
        return "benders"
    if "monolithic" in probe or "slackpenalty" in run_dir.name.lower():
        return "monolithic"
    return "unknown"


def _detect_dataset(run_dir: Path) -> str:
    metadata = _read_json(run_dir / "run_metadata.json")
    benders = _read_json(run_dir / "logs" / "benders_manifest.json")
    for value in (metadata.get("dataset"), benders.get("dataset")):
        if str(value).lower() in {"small", "full"}:
            return str(value).lower()
    match = re.search(r"(?:^|_)(small|full)(?:_|$)", run_dir.name, re.I)
    return match.group(1).lower() if match else "unknown"


def _detect_redirection(run_dir: Path) -> str:
    metadata = _read_json(run_dir / "run_metadata.json")
    benders = _read_json(run_dir / "logs" / "benders_manifest.json")
    for value in (metadata.get("scenario"), benders.get("scenario")):
        if value in {"with_redirection", "no_redirection"}:
            return str(value)
    lower = run_dir.name.lower()
    return "no_redirection" if "no_redirection" in lower else "with_redirection"


def _detect_technology(run_dir: Path) -> tuple[bool, bool, str]:
    lower = run_dir.name.lower()
    pv = "withpv" in lower or "_pv_" in lower
    bess = "withbess" in lower or "_bess_" in lower
    if "nopv" in lower:
        pv = False
    if "nobess" in lower:
        bess = False
    label = (
        "Chargers + PV + BESS" if pv and bess else
        "Chargers + PV" if pv else
        "Chargers + BESS" if bess else
        "Chargers only"
    )
    return pv, bess, label


def _auto_label(run_dir: Path) -> str:
    _, _, tech = _detect_technology(run_dir)
    redir = "With redirection" if _detect_redirection(run_dir) == "with_redirection" else "No redirection"
    return f"{tech} | {redir}"


def _runtime_seconds(run_dir: Path, method: str) -> float:
    scalars = read_scalar_file(run_dir)
    for key in ("total_runtime_seconds", "optimization_seconds", "decomposition_seconds"):
        try:
            value = float(scalars.get(key, math.nan))
            if math.isfinite(value):
                return value
        except Exception:
            pass
    if method == "lbbd":
        history = cr._read_csv_if_present(run_dir / "results" / "lbbd_history.csv")
    elif method == "benders":
        history = cr._read_csv_if_present(run_dir / "iterations" / "benders_iteration_history.csv")
    else:
        history = None
    if history is not None and "elapsed_seconds" in history.columns:
        values = pd.to_numeric(history["elapsed_seconds"], errors="coerce").dropna()
        if not values.empty:
            return float(values.max())
    return math.nan


def _scenario_summary(labels: list[str], metrics: list[dict[str, Any]]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for label, key, indent, kind in cr.ROW_SPEC:
        row: dict[str, Any] = {"Metric": label, "MetricKey": key or "", "Indent": indent, "Format": kind}
        if key is not None:
            for scenario_label, values in zip(labels, metrics):
                row[scenario_label] = values.get(key, math.nan)
        rows.append(row)
    return pd.DataFrame(rows)


def _changes_vs_baseline(labels: list[str], metrics: list[dict[str, Any]], baseline_idx: int) -> pd.DataFrame:
    base_label = labels[baseline_idx]
    base = metrics[baseline_idx]
    rows = []
    for label, key, _, kind in cr.ROW_SPEC:
        if key is None:
            continue
        row = {
            "Metric": label,
            "MetricKey": key,
            "Format": kind,
            f"Baseline ({base_label})": base.get(key, math.nan),
        }
        for j, scenario_label in enumerate(labels):
            if j == baseline_idx:
                continue
            value = metrics[j].get(key, math.nan)
            row[f"{scenario_label} value"] = value
            row[f"{scenario_label} - baseline"] = cr._difference(value, base.get(key, math.nan))
            row[f"{scenario_label} relative change"] = cr._relative(value, base.get(key, math.nan))
        rows.append(row)
    return pd.DataFrame(rows)


def _redirection_effects(labels: list[str], run_dirs: list[Path], metrics: list[dict[str, Any]]) -> pd.DataFrame:
    groups: dict[str, dict[str, int]] = {}
    for idx, run_dir in enumerate(run_dirs):
        _, _, tech = _detect_technology(run_dir)
        groups.setdefault(tech, {})[_detect_redirection(run_dir)] = idx
    rows = []
    for tech, pair in groups.items():
        if not {"no_redirection", "with_redirection"}.issubset(pair):
            continue
        no_i, yes_i = pair["no_redirection"], pair["with_redirection"]
        no, yes = metrics[no_i], metrics[yes_i]
        rows.append({
            "Technology": tech,
            "No-redirection run": labels[no_i],
            "With-redirection run": labels[yes_i],
            "Profit gain from redirection (SEK/yr)": cr._difference(yes.get("annual_profit_SEK"), no.get("annual_profit_SEK")),
            "Profit gain from redirection (%)": cr._relative(yes.get("annual_profit_SEK"), no.get("annual_profit_SEK")),
            "Grid electricity change (kWh/yr)": cr._difference(yes.get("grid_total_kWh"), no.get("grid_total_kWh")),
            "Public capacity change (kWh/30 min)": cr._difference(yes.get("public_charger_capacity_kWh_per_slot"), no.get("public_charger_capacity_kWh_per_slot")),
            "Redirected public demand (kWh/yr)": yes.get("energy_redirected_kWh", 0.0),
            "Redirected public-demand share": yes.get("share_redirected_public_demand", 0.0),
            "Redirection incentive (SEK/yr)": yes.get("redirection_total_cost_SEK", 0.0),
        })
    return pd.DataFrame(rows)


def _rankings(labels: list[str], metrics: list[dict[str, Any]], runtimes: list[float]) -> pd.DataFrame:
    rows = []
    for label, values, runtime in zip(labels, metrics, runtimes):
        rows.append({
            "Scenario": label,
            "Net profit (SEK/yr)": values.get("annual_profit_SEK", math.nan),
            "Total runtime (s)": runtime,
            "Grid electricity (kWh/yr)": values.get("grid_total_kWh", math.nan),
            "Public capacity (kWh/30 min)": values.get("public_charger_capacity_kWh_per_slot", math.nan),
            "Annual slack (kWh)": values.get("annual_slack_kWh", math.nan),
        })
    frame = pd.DataFrame(rows)
    if not frame.empty:
        frame["Profit rank"] = pd.to_numeric(frame["Net profit (SEK/yr)"], errors="coerce").rank(ascending=False, method="min")
        frame["Runtime rank"] = pd.to_numeric(frame["Total runtime (s)"], errors="coerce").rank(ascending=True, method="min")
    return frame


def _run_metadata(labels: list[str], run_dirs: list[Path], method: str, runtimes: list[float]) -> pd.DataFrame:
    rows = []
    for label, run_dir, runtime in zip(labels, run_dirs, runtimes):
        pv, bess, tech = _detect_technology(run_dir)
        rows.append({
            "Scenario": label,
            "Method": method.capitalize() if method != "lbbd" else "LBBD",
            "Dataset": _detect_dataset(run_dir),
            "Redirection": _detect_redirection(run_dir),
            "PV enabled": pv,
            "BESS enabled": bess,
            "Technology": tech,
            "Total runtime (s)": runtime,
            "Run folder": str(run_dir),
        })
    return pd.DataFrame(rows)


def _complexity_summary(labels: list[str], run_dirs: list[Path], method: str) -> pd.DataFrame:
    rows = []
    prefix = "initial_master" if method == "lbbd" else "master" if method == "benders" else "main_model"
    for label, run_dir in zip(labels, run_dirs):
        scalars = read_scalar_file(run_dir)
        detail = collect_solver_log_detail(run_dir)
        summary = summarize_solver_logs(detail)
        role = "master" if method in {"benders", "lbbd"} else "main_model"
        primary = summary[summary["Model role"] == role] if not summary.empty and "Model role" in summary.columns else pd.DataFrame()
        def sval(key: str) -> Any:
            return scalars.get(key, math.nan)
        def q(col: str) -> Any:
            if primary.empty or col not in primary.columns:
                return math.nan
            vals = pd.to_numeric(primary[col], errors="coerce").dropna()
            return float(vals.max()) if not vals.empty else math.nan
        rows.append({
            "Scenario": label,
            "Total runtime (s)": sval("total_runtime_seconds"),
            "Peak process-tree RSS (MB)": sval("peak_process_tree_rss_MB"),
            "Input loading (s)": sval("input_load_seconds"),
            "Preprocessing (s)": sval("preprocessing_seconds"),
            "Primary model build (s)": sval("master_build_seconds" if method == "lbbd" else "model_build_seconds"),
            "Primary variables": sval(f"{prefix}_variables_total"),
            "Primary binary variables": sval(f"{prefix}_variables_binary"),
            "Primary integer variables": sval(f"{prefix}_variables_integer_nonbinary"),
            "Primary continuous variables": sval(f"{prefix}_variables_continuous"),
            "Primary active constraints": sval(f"{prefix}_constraints_active"),
            "Solver initial rows": q("Max initial_rows"),
            "Solver initial columns": q("Max initial_columns"),
            "Solver presolved rows": q("Max presolved_rows"),
            "Solver presolved columns": q("Max presolved_columns"),
            "Solver nodes": q("Total nodes_explored"),
            "Solver work units": q("Total work_units"),
        })
    return pd.DataFrame(rows)


def _raw_metrics(labels: list[str], metrics: list[dict[str, Any]]) -> pd.DataFrame:
    keys = sorted(set().union(*(set(values) for values in metrics)))
    rows = []
    for key in keys:
        row = {"Metric": key}
        for label, values in zip(labels, metrics):
            row[label] = values.get(key, math.nan)
        rows.append(row)
    return pd.DataFrame(rows)


def _consistency_checks(run_dirs: list[Path], method: str, metrics: list[dict[str, Any]]) -> pd.DataFrame:
    detected_methods = [_detect_method(run_dir) for run_dir in run_dirs]
    datasets = [_detect_dataset(run_dir) for run_dir in run_dirs]
    checks = [
        ("At least two runs supplied", len(run_dirs) >= 2, str(len(run_dirs))),
        ("All runs match requested method", all(v == method for v in detected_methods), ", ".join(detected_methods)),
        ("All runs use the same dataset", len(set(datasets)) == 1, ", ".join(datasets)),
        ("Every run contains model_summary.csv", all((r / "results" / "model_summary.csv").exists() for r in run_dirs), ""),
    ]
    for idx, values in enumerate(metrics, start=1):
        slack = values.get("annual_slack_kWh", math.nan)
        checks.append((f"Run {idx}: annual slack available", isinstance(slack, (int, float)) and math.isfinite(float(slack)), str(slack)))
    return pd.DataFrame(checks, columns=["Check", "Pass", "Details"])


def _style_workbook(path: Path, summary: pd.DataFrame) -> None:
    wb = load_workbook(path)
    navy, light, pale, white = "1F4E78", "D9EAF7", "EEF5FA", "FFFFFF"
    thin = Side(style="thin", color="B7C9D6")
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.fill = PatternFill("solid", fgColor=navy)
                cell.font = Font(bold=True, color=white)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(bottom=thin)
        sheet.auto_filter.ref = sheet.dimensions
        for col in range(1, sheet.max_column + 1):
            width = max(len(str(sheet.cell(r, col).value or "")) for r in range(1, min(sheet.max_row, 100) + 1)) + 2
            sheet.column_dimensions[get_column_letter(col)].width = min(38, max(12, width))
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws = wb["Scenario summary"]
    ws.freeze_panes = "B2"
    for idx, (_, _, indent, kind) in enumerate(cr.ROW_SPEC, start=2):
        ws.cell(idx, 1).alignment = Alignment(indent=indent, vertical="center")
        if kind == "section":
            for col in range(1, ws.max_column + 1):
                ws.cell(idx, col).fill = PatternFill("solid", fgColor=light)
                ws.cell(idx, col).font = Font(bold=True, color=navy)
        elif indent:
            for col in range(1, ws.max_column + 1):
                ws.cell(idx, col).fill = PatternFill("solid", fgColor=pale)
        fmt = "#,##0" if kind == "count" else "0.0000%" if kind == "percent" else "#,##0.00"
        for col in range(5, ws.max_column + 1):
            ws.cell(idx, col).number_format = fmt if kind != "section" else "General"
    ws.column_dimensions["A"].width = 42
    # Hide technical helper columns but keep them in the workbook for reproducibility.
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].hidden = True

    for sheet_name in ("Changes vs baseline", "Redirection effects"):
        if sheet_name not in wb.sheetnames:
            continue
        sh = wb[sheet_name]
        for row in range(2, sh.max_row + 1):
            for col in range(1, sh.max_column + 1):
                header = str(sh.cell(1, col).value or "").lower()
                if "%" in header or "relative" in header or "share" in header:
                    sh.cell(row, col).number_format = "0.0000%"
                elif isinstance(sh.cell(row, col).value, (int, float)):
                    sh.cell(row, col).number_format = "#,##0.00"
    wb.save(path)


def main() -> int:
    args = parse_args()
    root = Path(args.project_root).resolve()
    if len(args.run) < 2:
        raise SystemExit("At least two --run arguments are required.")
    run_dirs = [_resolve(root, value) for value in args.run]
    for run_dir in run_dirs:
        if not run_dir.exists():
            raise FileNotFoundError(run_dir)
    method = args.method.lower()
    detected = [_detect_method(run_dir) for run_dir in run_dirs]
    if any(value != method for value in detected):
        raise ValueError(f"All runs must be {method}; detected methods: {detected}")
    datasets = [_detect_dataset(run_dir) for run_dir in run_dirs]
    if len(set(datasets)) != 1:
        raise ValueError(f"Scenario comparison requires one dataset; detected: {datasets}")
    if not (1 <= int(args.baseline_index) <= len(run_dirs)):
        raise ValueError("--baseline-index must be between 1 and the number of runs.")
    baseline_idx = int(args.baseline_index) - 1

    capacity = cr._load_capacity_config(root)
    metrics = [cr._enrich(cr._read_summary(run_dir), run_dir, capacity) for run_dir in run_dirs]
    if args.label:
        if len(args.label) != len(run_dirs):
            raise ValueError("When --label is used, provide one label for every --run.")
        labels = list(args.label)
    else:
        labels = [_auto_label(run_dir) for run_dir in run_dirs]
        # Ensure duplicate labels remain unambiguous.
        seen: dict[str, int] = {}
        for idx, label in enumerate(labels):
            seen[label] = seen.get(label, 0) + 1
            if seen[label] > 1:
                labels[idx] = f"{label} #{seen[label]}"

    runtimes = [_runtime_seconds(run_dir, method) for run_dir in run_dirs]
    summary = _scenario_summary(labels, metrics)
    changes = _changes_vs_baseline(labels, metrics, baseline_idx)
    redirection = _redirection_effects(labels, run_dirs, metrics)
    rankings = _rankings(labels, metrics, runtimes)
    metadata = _run_metadata(labels, run_dirs, method, runtimes)
    complexity = _complexity_summary(labels, run_dirs, method)
    raw = _raw_metrics(labels, metrics)
    checks = _consistency_checks(run_dirs, method, metrics)

    if args.out:
        output = _resolve(root, args.out)
    else:
        stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output = root / "runs" / "comparisons" / f"{stamp}_{method}_scenario_comparison.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Scenario summary", index=False)
        changes.to_excel(writer, sheet_name="Changes vs baseline", index=False)
        redirection.to_excel(writer, sheet_name="Redirection effects", index=False)
        rankings.to_excel(writer, sheet_name="Scenario ranking", index=False)
        complexity.to_excel(writer, sheet_name="Computational summary", index=False)
        metadata.to_excel(writer, sheet_name="Run metadata", index=False)
        raw.to_excel(writer, sheet_name="Raw metrics", index=False)
        checks.to_excel(writer, sheet_name="Consistency checks", index=False)
    _style_workbook(output, summary)
    print(f"Scenario comparison workbook written to: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
